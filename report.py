import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image

# Data load
df = pd.read_excel('data_sheet.xlsx')
total = df[df['Area'] == 'Total'].copy()

clean = total[['States/UTs',
               df.columns[7],
               df.columns[18],
               df.columns[13],
               df.columns[16]]].copy()

clean.columns = ['State', 'Sex_Ratio',
                 'Female_Literacy',
                 'Sanitation', 'Health_Insurance']

for col in clean.columns[1:]:
    clean[col] = pd.to_numeric(clean[col], errors='coerce')

states = clean[clean['State'] != 'India'].reset_index(drop=True)

# Excel banana shuru
wb = openpyxl.Workbook()

# ── Sheet 1: Clean Data ──
ws1 = wb.active
ws1.title = "Clean Data"

# Title row
ws1.merge_cells('A1:E1')
title = ws1['A1']
title.value = "NFHS-5 Healthcare Analysis — Clean Data"
title.font = Font(bold=True, size=14, color="FFFFFF")
title.fill = PatternFill("solid", fgColor="1A237E")
title.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

# Headers
headers = ['State', 'Sex Ratio', 'Female Literacy (%)',
           'Sanitation (%)', 'Health Insurance (%)']
header_colors = ['1565C0', '1565C0', '1565C0', '1565C0', '1565C0']

for j, (h, c) in enumerate(zip(headers, header_colors)):
    cell = ws1.cell(row=2, column=j+1, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=c)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Data rows
for r, (_, row) in enumerate(states.iterrows()):
    bg = "ECEFF1" if r % 2 == 0 else "FFFFFF"
    vals = [row['State'], row['Sex_Ratio'],
            row['Female_Literacy'], row['Sanitation'],
            row['Health_Insurance']]
    for j, val in enumerate(vals):
        cell = ws1.cell(row=r+3, column=j+1, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal='center' if j > 0 else 'left')
        # Color coding
        if j == 2:  # Female Literacy
            if isinstance(val, float):
                if val >= 80:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")
                elif val < 65:
                    cell.fill = PatternFill("solid", fgColor="FFCDD2")

# Column widths
ws1.column_dimensions['A'].width = 30
for col in ['B', 'C', 'D', 'E']:
    ws1.column_dimensions[col].width = 20

# ── Sheet 2: Charts ──
ws2 = wb.create_sheet("Charts")
ws2.merge_cells('A1:J1')
t2 = ws2['A1']
t2.value = "Healthcare Analysis — Charts"
t2.font = Font(bold=True, size=14, color="FFFFFF")
t2.fill = PatternFill("solid", fgColor="1A237E")
t2.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

# Chart image add karo
try:
    img = Image('chart1.png')
    img.width = 500
    img.height = 300
    ws2.add_image(img, 'A3')
    print("Chart image add ho gaya!")
except:
    print("Chart image nahi mili — pehle step2 run karo")

# ── Sheet 3: Recommendations ──
ws3 = wb.create_sheet("Recommendations")
ws3.merge_cells('A1:D1')
t3 = ws3['A1']
t3.value = "Key Findings & Recommendations"
t3.font = Font(bold=True, size=14, color="FFFFFF")
t3.fill = PatternFill("solid", fgColor="1A237E")
t3.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 35

# Headers
for j, h in enumerate(['Issue', 'Finding', 'Recommendation', 'Priority']):
    cell = ws3.cell(row=2, column=j+1, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="37474F")
    cell.alignment = Alignment(horizontal='center')

# Recommendations data
recs = [
    ("Female Literacy", "Bihar 57%, Rajasthan 60%",
     "Girl education campaigns ", "High"),
    ("Sanitation", "Bihar 49%",
     "Swachh Bharat Mission", "High"),
    ("Health Insurance", "National avg 41%",
     "Ayushman Bharat expand ", "Medium"),
    ("Sex Ratio", "in some states are less then 900 ",
     "Beti Bachao campaign strengthen ", "High"),
]

priority_colors = {"High": "FFCDD2", "Medium": "FFF9C4", "Low": "C8E6C9"}

for r, (issue, finding, rec, priority) in enumerate(recs):
    bg = priority_colors[priority]
    for j, val in enumerate([issue, finding, rec, priority]):
        cell = ws3.cell(row=r+3, column=j+1, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal='left', wrap_text=True)

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 35
ws3.column_dimensions['D'].width = 12

# Save 
wb.save('My_Healthcare_Report.xlsx')
print("\nReport ready! My_Healthcare_Report.xlsx are ready!")
print("Go to the folder  see the report!")